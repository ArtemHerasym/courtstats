from pathlib import Path
from app.calculations.basketball import (
    calculate_points,
    determine_game_result,
)
from openpyxl import load_workbook

from dataclasses import dataclass
from datetime import date

from app.models.game import VenueType
from openpyxl.worksheet.worksheet import Worksheet
from datetime import date, datetime

def _build_header_map(
    sheet: Worksheet,
) -> dict[str, int]:
    header_row = next(
        sheet.iter_rows(
            min_row=1,
            max_row=1,
            values_only=True,
        )
    )

    return {
        str(value).strip(): index
        for index, value in enumerate(header_row)
        if value is not None
    }

def _normalize_optional_text(
    value: object,
) -> str | None:
    if value is None:
        return None

    normalized = str(value).strip()

    if not normalized:
        return None

    return normalized


@dataclass(frozen=True)
class HistoricalRosterRow:
    full_name: str
    jersey_number: int | None
    position: str | None
    grade_level: str | None


@dataclass(frozen=True)
class HistoricalGameRow:
    workbook_game_id: str
    game_date: date
    opponent_name: str
    venue_type: VenueType
    opponent_score: int

    # Read only for pre-import verification.
    # These will never be stored in the database.
    workbook_team_score: int
    workbook_result: str


@dataclass(frozen=True)
class HistoricalPlayerStatsRow:
    workbook_game_id: str
    player_name: str

    three_point_attempts: int
    three_point_makes: int
    two_point_attempts: int
    two_point_makes: int
    free_throw_attempts: int
    free_throw_makes: int

    turnovers: int
    assists: int
    offensive_rebounds: int
    defensive_rebounds: int
    steals: int
    deflections: int
    personal_fouls: int


@dataclass(frozen=True)
class HistoricalWorkbookData:
    roster: list[HistoricalRosterRow]
    games: list[HistoricalGameRow]
    player_stats: list[HistoricalPlayerStatsRow]


def normalize_venue(value: str) -> VenueType:
    normalized = value.strip().upper()

    venue_mapping = {
        "HOME": VenueType.HOME,
        "AWAY": VenueType.AWAY,
        "AVAY": VenueType.AWAY,
        "NEUTRAL": VenueType.NEUTRAL,
    }

    try:
        return venue_mapping[normalized]
    except KeyError as exc:
        raise ValueError(
            f"Unknown historical venue value: {value!r}"
        ) from exc



def parse_jersey_number(
    value: str | int | float | None,
) -> int | None:
    if value is None:
        return None

    if isinstance(value, bool):
        raise ValueError(
            f"Invalid jersey number: {value!r}"
        )

    if isinstance(value, int):
        jersey_number = value

    elif isinstance(value, float):
        if not value.is_integer():
            raise ValueError(
                f"Invalid jersey number: {value!r}"
            )

        jersey_number = int(value)

    elif isinstance(value, str):
        normalized = value.strip()

        if not normalized:
            return None

        if normalized.startswith("#"):
            normalized = normalized[1:].strip()

        if not normalized.isdigit():
            raise ValueError(
                f"Invalid jersey number: {value!r}"
            )

        jersey_number = int(normalized)

    else:
        raise ValueError(
            f"Invalid jersey number: {value!r}"
        )

    if jersey_number < 0:
        raise ValueError(
            f"Jersey number cannot be negative: {jersey_number}"
        )

    return jersey_number


def parse_roster_sheet(
    sheet: Worksheet,
) -> list[HistoricalRosterRow]:
    headers = _build_header_map(sheet)

    required_headers = {
        "Player Name",
        "Jersey Number",
        "Position",
        "Grade",
    }

    missing_headers = required_headers - headers.keys()

    if missing_headers:
        raise ValueError(
            "Team Roster sheet is missing required headers: "
            + ", ".join(sorted(missing_headers))
        )

    roster_rows: list[HistoricalRosterRow] = []

    for row in sheet.iter_rows(
        min_row=2,
        values_only=True,
    ):
        player_name_value = row[
            headers["Player Name"]
        ]

        # Completely blank roster rows are ignored.
        if player_name_value is None:
            if all(value is None for value in row):
                continue

            raise ValueError(
                "Team Roster row contains data but has no Player Name"
            )

        player_name = str(player_name_value).strip()

        if not player_name:
            raise ValueError(
                "Team Roster contains a blank Player Name"
            )

        roster_rows.append(
            HistoricalRosterRow(
                full_name=player_name,
                jersey_number=parse_jersey_number(
                    row[headers["Jersey Number"]]
                ),
                position=_normalize_optional_text(
                    row[headers["Position"]]
                ),
                grade_level=_normalize_optional_text(
                    row[headers["Grade"]]
                ),
            )
        )

    return roster_rows


def _parse_nonnegative_int(
    value: object,
    field_name: str,
) -> int:
    if isinstance(value, bool):
        raise ValueError(
            f"{field_name} must be a nonnegative integer: {value!r}"
        )

    if isinstance(value, int):
        result = value

    elif isinstance(value, float):
        if not value.is_integer():
            raise ValueError(
                f"{field_name} must be a nonnegative integer: {value!r}"
            )
        result = int(value)

    elif isinstance(value, str):
        normalized = value.strip()

        if not normalized.isdigit():
            raise ValueError(
                f"{field_name} must be a nonnegative integer: {value!r}"
            )

        result = int(normalized)

    else:
        raise ValueError(
            f"{field_name} must be a nonnegative integer: {value!r}"
        )

    if result < 0:
        raise ValueError(
            f"{field_name} cannot be negative: {result}"
        )

    return result

def _parse_game_date(
    value: object,
) -> date:
    if isinstance(value, datetime):
        return value.date()

    if isinstance(value, date):
        return value

    if isinstance(value, str):
        normalized = value.strip()

        for date_format in (
            "%m/%d/%Y",
            "%Y-%m-%d",
        ):
            try:
                return datetime.strptime(
                    normalized,
                    date_format,
                ).date()
            except ValueError:
                continue

    raise ValueError(
        f"Invalid historical game date: {value!r}"
    )


def parse_games_sheet(
    sheet: Worksheet,
) -> list[HistoricalGameRow]:
    headers = _build_header_map(sheet)

    required_headers = {
        "Game ID",
        "Date",
        "Opponent",
        "Home/Away",
        "Team score",
        "Opp. Score",
        "Result",
    }

    missing_headers = required_headers - headers.keys()

    if missing_headers:
        raise ValueError(
            "Games sheet is missing required headers: "
            + ", ".join(sorted(missing_headers))
        )

    games: list[HistoricalGameRow] = []
    seen_game_ids: set[str] = set()

    for row in sheet.iter_rows(
        min_row=2,
        values_only=True,
    ):
        if all(value is None for value in row):
            continue

        game_id_value = row[headers["Game ID"]]

        if game_id_value is None:
            raise ValueError(
                "Games row contains data but has no Game ID"
            )

        workbook_game_id = str(
            game_id_value
        ).strip()

        if not workbook_game_id:
            raise ValueError(
                "Games sheet contains a blank Game ID"
            )

        if workbook_game_id in seen_game_ids:
            raise ValueError(
                f"Duplicate historical Game ID: {workbook_game_id}"
            )

        seen_game_ids.add(workbook_game_id)

        opponent_value = row[
            headers["Opponent"]
        ]

        if opponent_value is None:
            raise ValueError(
                f"{workbook_game_id} has no opponent"
            )

        opponent_name = str(
            opponent_value
        ).strip()

        if not opponent_name:
            raise ValueError(
                f"{workbook_game_id} has a blank opponent"
            )

        result_value = row[
            headers["Result"]
        ]

        if result_value is None:
            raise ValueError(
                f"{workbook_game_id} has no result"
            )

        workbook_result = str(
            result_value
        ).strip().upper()

        if workbook_result not in {
            "WIN",
            "LOSS",
            "TIE",
        }:
            raise ValueError(
                f"{workbook_game_id} has invalid result: "
                f"{result_value!r}"
            )

        venue_value = row[
            headers["Home/Away"]
        ]

        if venue_value is None:
            raise ValueError(
                f"{workbook_game_id} has no venue"
            )

        games.append(
            HistoricalGameRow(
                workbook_game_id=workbook_game_id,
                game_date=_parse_game_date(
                    row[headers["Date"]]
                ),
                opponent_name=opponent_name,
                venue_type=normalize_venue(
                    str(venue_value)
                ),
                opponent_score=_parse_nonnegative_int(
                    row[headers["Opp. Score"]],
                    "Opponent score",
                ),
                workbook_team_score=_parse_nonnegative_int(
                    row[headers["Team score"]],
                    "Team score",
                ),
                workbook_result=workbook_result,
            )
        )

    return games


def parse_player_stats_sheet(
    sheet: Worksheet,
) -> list[HistoricalPlayerStatsRow]:
    headers = _build_header_map(sheet)

    required_headers = {
        "Game Id",
        "Player Name",
        "3PTA",
        "3PTM",
        "2PTA",
        "2PTM",
        "FTA",
        "FTM",
        "TO",
        "Assists",
        "Off. Reb.",
        "Def. Reb.",
        "Steals",
        "Defl.",
        "Pers. Fouls",
    }

    missing_headers = required_headers - headers.keys()

    if missing_headers:
        raise ValueError(
            "Player Stats sheet is missing required headers: "
            + ", ".join(sorted(missing_headers))
        )

    stats_rows: list[HistoricalPlayerStatsRow] = []
    seen_player_games: set[tuple[str, str]] = set()

    for row in sheet.iter_rows(
        min_row=2,
        values_only=True,
    ):
        if all(value is None for value in row):
            continue

        game_id_value = row[
            headers["Game Id"]
        ]

        if game_id_value is None:
            raise ValueError(
                "Player Stats row contains data but has no Game Id"
            )

        workbook_game_id = str(
            game_id_value
        ).strip()

        if not workbook_game_id:
            raise ValueError(
                "Player Stats contains a blank Game Id"
            )

        player_name_value = row[
            headers["Player Name"]
        ]

        if player_name_value is None:
            raise ValueError(
                f"{workbook_game_id} contains a stats row "
                "with no Player Name"
            )

        player_name = str(
            player_name_value
        ).strip()

        if not player_name:
            raise ValueError(
                f"{workbook_game_id} contains a stats row "
                "with a blank Player Name"
            )

        # Spreadsheet aggregate rows are calculated data,
        # not PlayerGameStats records.
        if player_name.casefold() == "team total":
            continue

        player_game_key = (
            workbook_game_id,
            player_name.casefold(),
        )

        if player_game_key in seen_player_games:
            raise ValueError(
                "Duplicate historical player stats row: "
                f"{workbook_game_id} / {player_name}"
            )

        seen_player_games.add(
            player_game_key
        )

        three_point_attempts = _parse_nonnegative_int(
            row[headers["3PTA"]],
            "3PTA",
        )
        three_point_makes = _parse_nonnegative_int(
            row[headers["3PTM"]],
            "3PTM",
        )
        two_point_attempts = _parse_nonnegative_int(
            row[headers["2PTA"]],
            "2PTA",
        )
        two_point_makes = _parse_nonnegative_int(
            row[headers["2PTM"]],
            "2PTM",
        )
        free_throw_attempts = _parse_nonnegative_int(
            row[headers["FTA"]],
            "FTA",
        )
        free_throw_makes = _parse_nonnegative_int(
            row[headers["FTM"]],
            "FTM",
        )

        if three_point_makes > three_point_attempts:
            raise ValueError(
                f"{workbook_game_id} / {player_name}: "
                "3PTM cannot exceed 3PTA"
            )

        if two_point_makes > two_point_attempts:
            raise ValueError(
                f"{workbook_game_id} / {player_name}: "
                "2PTM cannot exceed 2PTA"
            )

        if free_throw_makes > free_throw_attempts:
            raise ValueError(
                f"{workbook_game_id} / {player_name}: "
                "FTM cannot exceed FTA"
            )

        stats_rows.append(
            HistoricalPlayerStatsRow(
                workbook_game_id=workbook_game_id,
                player_name=player_name,
                three_point_attempts=three_point_attempts,
                three_point_makes=three_point_makes,
                two_point_attempts=two_point_attempts,
                two_point_makes=two_point_makes,
                free_throw_attempts=free_throw_attempts,
                free_throw_makes=free_throw_makes,
                turnovers=_parse_nonnegative_int(
                    row[headers["TO"]],
                    "TO",
                ),
                assists=_parse_nonnegative_int(
                    row[headers["Assists"]],
                    "Assists",
                ),
                offensive_rebounds=_parse_nonnegative_int(
                    row[headers["Off. Reb."]],
                    "Off. Reb.",
                ),
                defensive_rebounds=_parse_nonnegative_int(
                    row[headers["Def. Reb."]],
                    "Def. Reb.",
                ),
                steals=_parse_nonnegative_int(
                    row[headers["Steals"]],
                    "Steals",
                ),
                deflections=_parse_nonnegative_int(
                    row[headers["Defl."]],
                    "Defl.",
                ),
                personal_fouls=_parse_nonnegative_int(
                    row[headers["Pers. Fouls"]],
                    "Pers. Fouls",
                ),
            )
        )

    return stats_rows

REQUIRED_SHEETS = {
    "Games",
    "Player Stats",
    "Team Roster",
}

def parse_historical_workbook(
    workbook_path: str | Path,
) -> HistoricalWorkbookData:
    workbook = load_workbook(
        filename=workbook_path,
        read_only=True,
        data_only=True,
    )

    try:
        missing_sheets = REQUIRED_SHEETS - set(
            workbook.sheetnames
        )

        if missing_sheets:
            raise ValueError(
                "Historical workbook is missing required sheets: "
                + ", ".join(sorted(missing_sheets))
            )

        roster = parse_roster_sheet(
            workbook["Team Roster"]
        )

        games = parse_games_sheet(
            workbook["Games"]
        )

        player_stats = parse_player_stats_sheet(
            workbook["Player Stats"]
        )

        return HistoricalWorkbookData(
            roster=roster,
            games=games,
            player_stats=player_stats,
        )

    finally:
        workbook.close()

def validate_historical_workbook(
    data: HistoricalWorkbookData,
) -> None:
    if len(data.games) != 15:
        raise ValueError(
            "Historical workbook must contain exactly 15 games; "
            f"found {len(data.games)}"
        )

    if not data.roster:
        raise ValueError(
            "Historical workbook roster cannot be empty"
        )

    game_ids = {
        game.workbook_game_id
        for game in data.games
    }

    roster_names = {
        player.full_name.casefold()
        for player in data.roster
    }

    if len(roster_names) != len(data.roster):
        raise ValueError(
            "Historical roster contains duplicate player names"
        )

    jersey_numbers = [
        player.jersey_number
        for player in data.roster
        if player.jersey_number is not None
    ]

    if len(jersey_numbers) != len(set(jersey_numbers)):
        raise ValueError(
            "Historical roster contains duplicate jersey numbers"
        )

    for stats in data.player_stats:
        if stats.workbook_game_id not in game_ids:
            raise ValueError(
                "Player Stats references unknown Game ID: "
                f"{stats.workbook_game_id}"
            )

        if stats.player_name.casefold() not in roster_names:
            raise ValueError(
                "Player Stats references player not found "
                f"in Team Roster: {stats.player_name}"
            )

        stats_by_game: dict[
            str,
            list[HistoricalPlayerStatsRow],
        ] = {
            game_id: []
            for game_id in game_ids
        }

        for stats in data.player_stats:
            stats_by_game[
                stats.workbook_game_id
            ].append(stats)

        for game in data.games:
            game_stats = stats_by_game[
                game.workbook_game_id
            ]

            if not game_stats:
                raise ValueError(
                    f"{game.workbook_game_id} has no player statistics"
                )

            calculated_team_score = sum(
                calculate_points(
                    stats.two_point_makes,
                    stats.three_point_makes,
                    stats.free_throw_makes,
                )
                for stats in game_stats
            )

            if calculated_team_score != game.workbook_team_score:
                raise ValueError(
                    f"{game.workbook_game_id} team score mismatch: "
                    f"raw stats calculate to {calculated_team_score}, "
                    f"workbook says {game.workbook_team_score}"
                )

            calculated_result = determine_game_result(
                calculated_team_score,
                game.opponent_score,
            )

            if calculated_result != game.workbook_result:
                raise ValueError(
                    f"{game.workbook_game_id} result mismatch: "
                    f"raw stats calculate to {calculated_result}, "
                    f"workbook says {game.workbook_result}"
                )


def load_and_validate_historical_workbook(
    workbook_path: str | Path,
) -> HistoricalWorkbookData:
    data = parse_historical_workbook(
        workbook_path
    )

    validate_historical_workbook(data)

    return data


