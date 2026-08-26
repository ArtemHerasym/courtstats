import pytest
from sqlalchemy import select
from openpyxl import Workbook
from datetime import date, datetime
from sqlalchemy import select

from app.models.player import Player
from app.models.season import Season, SeasonStatus
from app.models.team import Team

from app.models.player_game_stats import (
    ParticipationStatus,
    PlayerGameStats,
)
from app.models.season_roster import SeasonRoster
from app.importers.historical_database import (
    HistoricalImportAlreadyExistsError,
    ensure_historical_season_not_imported,
    get_or_create_player,
    get_or_create_team,
    import_historical_data,
)
from app.models.season import Season

from app.importers.historical_workbook import (
    HistoricalGameRow,
    HistoricalPlayerStatsRow,
    HistoricalRosterRow,
    HistoricalWorkbookData,
    load_and_validate_historical_workbook,
    normalize_venue,
    parse_games_sheet,
    parse_historical_workbook,
    parse_jersey_number,
    parse_player_stats_sheet,
    parse_roster_sheet,
    validate_historical_workbook,
)
from app.models.game import Game, GameStatus, VenueType

def _create_player_stats_sheet():
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Player Stats"

    sheet.append(
        [
            "Game Id",
            "Player Name",
            "3PTA",
            "3PTM",
            "2PTA",
            "2PTM",
            "FTA",
            "FTM",
            "Total Points",
            "TO",
            "Assists",
            "Off. Reb.",
            "Def. Reb.",
            "Steals",
            "Defl.",
            "Pers. Fouls",
            "FG%",
            "TS%",
            "A:TO",
            "Total Reb",
        ]
    )

    return sheet


@pytest.mark.parametrize(
    ("value", "expected"),
    [
        ("Home", VenueType.HOME),
        ("Away", VenueType.AWAY),
        ("Avay", VenueType.AWAY),
        ("Neutral", VenueType.NEUTRAL),
        (" away ", VenueType.AWAY),
    ],
)
def test_normalize_venue(
    value,
    expected,
):
    assert normalize_venue(value) == expected


def test_normalize_venue_rejects_unknown_value():
    with pytest.raises(
        ValueError,
        match="Unknown historical venue value",
    ):
        normalize_venue("Somewhere")


@pytest.mark.parametrize(
    ("value", "expected"),
    [
        ("#2", 2),
        ("#11", 11),
        ("15", 15),
        (7, 7),
        (7.0, 7),
        ("", None),
        (None, None),
    ],
)
def test_parse_jersey_number(
    value,
    expected,
):
    assert parse_jersey_number(value) == expected


@pytest.mark.parametrize(
    "value",
    [
        "#ABC",
        "2.5",
        -1,
        3.5,
        True,
    ],
)
def test_parse_jersey_number_rejects_invalid_value(
    value,
):
    with pytest.raises(ValueError):
        parse_jersey_number(value)


def _create_roster_sheet():
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Team Roster"

    sheet.append(
        [
            "Player Name",
            "Jersey Number",
            "Position",
            "Grade",
        ]
    )

    return sheet


def test_parse_roster_sheet_returns_roster_rows():
    sheet = _create_roster_sheet()

    sheet.append(
        [
            "Kevin Lans",
            "#2",
            "PG",
            "12",
        ]
    )

    sheet.append(
        [
            "John Smith",
            "#11",
            "SG",
            "11",
        ]
    )

    result = parse_roster_sheet(sheet)

    assert len(result) == 2

    assert result[0].full_name == "Kevin Lans"
    assert result[0].jersey_number == 2
    assert result[0].position == "PG"
    assert result[0].grade_level == "12"

    assert result[1].full_name == "John Smith"
    assert result[1].jersey_number == 11
    assert result[1].position == "SG"
    assert result[1].grade_level == "11"


def test_parse_roster_sheet_skips_blank_rows():
    sheet = _create_roster_sheet()

    sheet.append(
        [
            "Kevin Lans",
            "#2",
            "PG",
            "12",
        ]
    )

    sheet.append(
        [
            None,
            None,
            None,
            None,
        ]
    )

    result = parse_roster_sheet(sheet)

    assert len(result) == 1


def test_parse_roster_sheet_rejects_missing_required_header():
    workbook = Workbook()
    sheet = workbook.active

    sheet.append(
        [
            "Player Name",
            "Jersey Number",
            "Position",
        ]
    )

    with pytest.raises(
        ValueError,
        match="missing required headers",
    ):
        parse_roster_sheet(sheet)

def test_parse_roster_sheet_rejects_row_without_player_name():
    sheet = _create_roster_sheet()

    sheet.append(
        [
            None,
            "#7",
            "PG",
            "12",
        ]
    )

    with pytest.raises(
        ValueError,
        match="has no Player Name",
    ):
        parse_roster_sheet(sheet)


def _create_games_sheet():
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Games"

    sheet.append(
        [
            "Game ID",
            "Date",
            "Opponent",
            "Home/Away",
            "Team score",
            "Opp. Score",
            "Result",
        ]
    )

    return sheet


def test_parse_games_sheet_returns_game_rows():
    sheet = _create_games_sheet()

    sheet.append(
        [
            "G001",
            datetime(2025, 11, 1),
            "FCP",
            "Home",
            70,
            60,
            "Win",
        ]
    )

    sheet.append(
        [
            "G002",
            datetime(2025, 11, 2),
            "DME",
            "Avay",
            55,
            65,
            "Loss",
        ]
    )

    result = parse_games_sheet(sheet)

    assert len(result) == 2

    assert result[0].workbook_game_id == "G001"
    assert result[0].game_date == date(2025, 11, 1)
    assert result[0].opponent_name == "FCP"
    assert result[0].venue_type == VenueType.HOME
    assert result[0].opponent_score == 60
    assert result[0].workbook_team_score == 70
    assert result[0].workbook_result == "WIN"

    assert result[1].workbook_game_id == "G002"
    assert result[1].venue_type == VenueType.AWAY
    assert result[1].workbook_result == "LOSS"


def test_parse_games_sheet_rejects_duplicate_game_id():
    sheet = _create_games_sheet()

    for _ in range(2):
        sheet.append(
            [
                "G001",
                datetime(2025, 11, 1),
                "FCP",
                "Home",
                70,
                60,
                "Win",
            ]
        )

    with pytest.raises(
        ValueError,
        match="Duplicate historical Game ID",
    ):
        parse_games_sheet(sheet)

def test_parse_games_sheet_rejects_missing_required_header():
   workbook = Workbook()
   sheet = workbook.active

   sheet.append(
       [
           "Game ID",
           "Date",
           "Opponent",
       ]
   )

   with pytest.raises(
           ValueError,
           match="Games sheet is missing required headers",
   ):
       parse_games_sheet(sheet)


def test_parse_games_sheet_rejects_invalid_result():
    sheet = _create_games_sheet()

    sheet.append(
        [
            "G001",
            datetime(2025, 11, 1),
            "FCP",
            "Home",
            70,
            60,
            "Maybe",
        ]
    )

    with pytest.raises(
        ValueError,
        match="invalid result",
    ):
        parse_games_sheet(sheet)


def test_parse_games_sheet_skips_blank_rows():
    sheet = _create_games_sheet()

    sheet.append(
        [
            "G001",
            datetime(2025, 11, 1),
            "FCP",
            "Home",
            70,
            60,
            "Win",
        ]
    )

    sheet.append(
        [None] * 7
    )

    result = parse_games_sheet(sheet)

    assert len(result) == 1


def test_parse_player_stats_sheet_returns_raw_stats():
    sheet = _create_player_stats_sheet()

    sheet.append(
        [
            "G001",
            "Kevin Lans",
            6,
            2,
            3,
            0,
            2,
            1,
            999,       # deliberately incorrect calculated value
            4,
            1,
            0,
            2,
            1,
            1,
            2,
            "wrong",
            "wrong",
            "wrong",
            "wrong",
        ]
    )

    result = parse_player_stats_sheet(sheet)

    assert len(result) == 1

    stats = result[0]

    assert stats.workbook_game_id == "G001"
    assert stats.player_name == "Kevin Lans"

    assert stats.three_point_attempts == 6
    assert stats.three_point_makes == 2
    assert stats.two_point_attempts == 3
    assert stats.two_point_makes == 0
    assert stats.free_throw_attempts == 2
    assert stats.free_throw_makes == 1

    assert stats.turnovers == 4
    assert stats.assists == 1
    assert stats.offensive_rebounds == 0
    assert stats.defensive_rebounds == 2
    assert stats.steals == 1
    assert stats.deflections == 1
    assert stats.personal_fouls == 2

def test_parse_player_stats_sheet_skips_team_total_rows():
    sheet = _create_player_stats_sheet()

    sheet.append(
        [
            "G001",
            "Kevin Lans",
            1, 1, 1, 1, 0, 0,
            4,
            0, 0, 0, 0, 0, 0, 0,
            None, None, None, None,
        ]
    )

    sheet.append(
        [
            "G001",
            "Team Total",
            20, 10, 30, 15, 10, 8,
            68,
            15, 20, 10, 20, 5, 5, 10,
            None, None, None, None,
        ]
    )

    result = parse_player_stats_sheet(sheet)

    assert len(result) == 1
    assert result[0].player_name == "Kevin Lans"

def test_parse_player_stats_sheet_keeps_zero_stat_player_row():
    sheet = _create_player_stats_sheet()

    sheet.append(
        [
            "G001",
            "Kevin Lans",
            0, 0, 0, 0, 0, 0,
            0,
            0, 0, 0, 0, 0, 0, 0,
            None, None, None, None,
        ]
    )

    result = parse_player_stats_sheet(sheet)

    assert len(result) == 1

    stats = result[0]

    assert stats.player_name == "Kevin Lans"
    assert stats.three_point_attempts == 0
    assert stats.two_point_attempts == 0
    assert stats.free_throw_attempts == 0


def test_parse_player_stats_sheet_rejects_duplicate_player_game_row():
    sheet = _create_player_stats_sheet()

    row = [
        "G001",
        "Kevin Lans",
        1, 0, 1, 0, 0, 0,
        0,
        0, 0, 0, 0, 0, 0, 0,
        None, None, None, None,
    ]

    sheet.append(row)
    sheet.append(row)

    with pytest.raises(
        ValueError,
        match="Duplicate historical player stats row",
    ):
        parse_player_stats_sheet(sheet)


def test_parse_player_stats_sheet_rejects_negative_raw_stat():
    sheet = _create_player_stats_sheet()

    sheet.append(
        [
            "G001",
            "Kevin Lans",
            -1, 0, 0, 0, 0, 0,
            0,
            0, 0, 0, 0, 0, 0, 0,
            None, None, None, None,
        ]
    )

    with pytest.raises(
        ValueError,
        match="3PTA",
    ):
        parse_player_stats_sheet(sheet)


def test_parse_player_stats_sheet_rejects_makes_above_attempts():
    sheet = _create_player_stats_sheet()

    sheet.append(
        [
            "G001",
            "Kevin Lans",
            2, 3, 0, 0, 0, 0,
            9,
            0, 0, 0, 0, 0, 0, 0,
            None, None, None, None,
        ]
    )

    with pytest.raises(
        ValueError,
        match="3PTM cannot exceed 3PTA",
    ):
        parse_player_stats_sheet(sheet)


def test_parse_player_stats_sheet_rejects_missing_required_header():
    workbook = Workbook()
    sheet = workbook.active

    sheet.append(
        [
            "Game Id",
            "Player Name",
            "3PTA",
        ]
    )

    with pytest.raises(
        ValueError,
        match="Player Stats sheet is missing required headers",
    ):
        parse_player_stats_sheet(sheet)

def test_parse_historical_workbook_rejects_missing_sheet(
    tmp_path,
):
    workbook = Workbook()

    games = workbook.active
    games.title = "Games"

    workbook.create_sheet("Team Roster")

    path = tmp_path / "missing-sheet.xlsx"
    workbook.save(path)

    with pytest.raises(
        ValueError,
        match="missing required sheets",
    ):
        parse_historical_workbook(path)


def test_validate_historical_workbook_rejects_unknown_game_id():
    data = HistoricalWorkbookData(
        roster=[
            HistoricalRosterRow(
                full_name="Kevin Lans",
                jersey_number=2,
                position="Guard",
                grade_level="11",
            )
        ],
        games=[
            HistoricalGameRow(
                workbook_game_id=f"G{i:03}",
                game_date=date(2025, 11, 1),
                opponent_name="Opponent",
                venue_type=VenueType.HOME,
                opponent_score=10,
                workbook_team_score=0,
                workbook_result="LOSS",
            )
            for i in range(1, 16)
        ],
        player_stats=[
            HistoricalPlayerStatsRow(
                workbook_game_id="G999",
                player_name="Kevin Lans",
                three_point_attempts=0,
                three_point_makes=0,
                two_point_attempts=0,
                two_point_makes=0,
                free_throw_attempts=0,
                free_throw_makes=0,
                turnovers=0,
                assists=0,
                offensive_rebounds=0,
                defensive_rebounds=0,
                steals=0,
                deflections=0,
                personal_fouls=0,
            )
        ],
    )

    with pytest.raises(
        ValueError,
        match="unknown Game ID",
    ):
        validate_historical_workbook(data)

def test_validate_historical_workbook_rejects_unknown_player():
    games = [
        HistoricalGameRow(
            workbook_game_id=f"G{i:03}",
            game_date=date(2025, 11, i),
            opponent_name="Opponent",
            venue_type=VenueType.HOME,
            opponent_score=10,
            workbook_team_score=0,
            workbook_result="LOSS",
        )
        for i in range(1, 16)
    ]

    data = HistoricalWorkbookData(
        roster=[
            HistoricalRosterRow(
                full_name="Kevin Lans",
                jersey_number=2,
                position="Guard",
                grade_level="11",
            )
        ],
        games=games,
        player_stats=[
            HistoricalPlayerStatsRow(
                workbook_game_id="G001",
                player_name="Unknown Player",
                three_point_attempts=0,
                three_point_makes=0,
                two_point_attempts=0,
                two_point_makes=0,
                free_throw_attempts=0,
                free_throw_makes=0,
                turnovers=0,
                assists=0,
                offensive_rebounds=0,
                defensive_rebounds=0,
                steals=0,
                deflections=0,
                personal_fouls=0,
            )
        ],
    )

    with pytest.raises(
        ValueError,
        match="not found in Team Roster",
    ):
        validate_historical_workbook(data)

def test_validate_historical_workbook_rejects_team_score_mismatch():
    games = []

    for i in range(1, 16):
        workbook_game_id = f"G{i:03}"

        if workbook_game_id == "G001":
            workbook_team_score = 10
            workbook_result = "TIE"
        else:
            workbook_team_score = 0
            workbook_result = "LOSS"

        games.append(
            HistoricalGameRow(
                workbook_game_id=workbook_game_id,
                game_date=date(2025, 11, i),
                opponent_name="Opponent",
                venue_type=VenueType.HOME,
                opponent_score=10,
                workbook_team_score=workbook_team_score,
                workbook_result=workbook_result,
            )
        )

    player_stats = []

    for i in range(1, 16):
        workbook_game_id = f"G{i:03}"

        if workbook_game_id == "G001":
            two_point_attempts = 1
            two_point_makes = 1
        else:
            two_point_attempts = 0
            two_point_makes = 0

        player_stats.append(
            HistoricalPlayerStatsRow(
                workbook_game_id=workbook_game_id,
                player_name="Kevin Lans",
                three_point_attempts=0,
                three_point_makes=0,
                two_point_attempts=two_point_attempts,
                two_point_makes=two_point_makes,
                free_throw_attempts=0,
                free_throw_makes=0,
                turnovers=0,
                assists=0,
                offensive_rebounds=0,
                defensive_rebounds=0,
                steals=0,
                deflections=0,
                personal_fouls=0,
            )
        )

    data = HistoricalWorkbookData(
        roster=[
            HistoricalRosterRow(
                full_name="Kevin Lans",
                jersey_number=2,
                position="Guard",
                grade_level="11",
            )
        ],
        games=games,
        player_stats=player_stats,
    )

    with pytest.raises(
        ValueError,
        match="team score mismatch",
    ):
        validate_historical_workbook(data)


def test_get_or_create_team_reuses_existing_team_case_insensitively(
    db_session,
):
    existing = Team(
        name="Jordan Christian Preparatory",
        abbreviation="JCP",
    )

    db_session.add(existing)
    db_session.commit()
    db_session.refresh(existing)

    result = get_or_create_team(
        db_session,
        "jordan christian preparatory",
        "JCP",
    )

    assert result.id == existing.id

    teams = list(
        db_session.scalars(
            select(Team)
        ).all()
    )

    assert len(teams) == 1


def test_get_or_create_player_reuses_existing_player_case_insensitively(
    db_session,
):
    existing = Player(
        full_name="Kevin Lans",
    )

    db_session.add(existing)
    db_session.commit()
    db_session.refresh(existing)

    result = get_or_create_player(
        db_session,
        "kevin lans",
    )

    assert result.id == existing.id

    players = list(
        db_session.scalars(
            select(Player)
        ).all()
    )

    assert len(players) == 1


def test_duplicate_historical_season_is_rejected(
    db_session,
):
    team = Team(
        name="Jordan Christian Preparatory",
        abbreviation="JCP",
    )

    db_session.add(team)
    db_session.commit()
    db_session.refresh(team)

    season = Season(
        team_id=team.id,
        name="2025-26",
    )

    db_session.add(season)
    db_session.commit()

    with pytest.raises(
        HistoricalImportAlreadyExistsError,
        match="already exists",
    ):
        ensure_historical_season_not_imported(
            db_session,
            team.id,
        )

def test_import_historical_data_creates_complete_historical_season(
    db_session,
):
    games = [
        HistoricalGameRow(
            workbook_game_id=f"G{i:03}",
            game_date=date(2025, 11, i),
            opponent_name="Historical Opponent",
            venue_type=VenueType.HOME,
            opponent_score=0,
            workbook_team_score=2,
            workbook_result="WIN",
        )
        for i in range(1, 16)
    ]

    player_stats = [
        HistoricalPlayerStatsRow(
            workbook_game_id=f"G{i:03}",
            player_name="Kevin Lans",
            three_point_attempts=0,
            three_point_makes=0,
            two_point_attempts=1,
            two_point_makes=1,
            free_throw_attempts=0,
            free_throw_makes=0,
            turnovers=0,
            assists=0,
            offensive_rebounds=0,
            defensive_rebounds=0,
            steals=0,
            deflections=0,
            personal_fouls=0,
        )
        for i in range(1, 16)
    ]

    data = HistoricalWorkbookData(
        roster=[
            HistoricalRosterRow(
                full_name="Kevin Lans",
                jersey_number=2,
                position="Guard",
                grade_level="11",
            )
        ],
        games=games,
        player_stats=player_stats,
    )

    season = import_historical_data(
        db_session,
        data,
    )

    assert season.name == "2025-26"
    assert season.status == SeasonStatus.COMPLETED
    assert season.start_date == date(2025, 11, 1)
    assert season.end_date == date(2025, 11, 15)

    roster_entries = list(
        db_session.scalars(
            select(SeasonRoster)
        ).all()
    )

    imported_games = list(
        db_session.scalars(
            select(Game)
        ).all()
    )

    imported_stats = list(
        db_session.scalars(
            select(PlayerGameStats)
        ).all()
    )

    teams = list(
        db_session.scalars(
            select(Team)
        ).all()
    )

    assert len(roster_entries) == 1

    assert len(imported_games) == 15
    assert all(
        game.status == GameStatus.COMPLETED
        for game in imported_games
    )

    assert len(imported_stats) == 15
    assert all(
        stats.participation_status
        == ParticipationStatus.PLAYED
        for stats in imported_stats
    )

    # JCP + one reused opponent.
    assert len(teams) == 2


def test_import_historical_data_rejects_second_import(
    db_session,
):
    games = [
        HistoricalGameRow(
            workbook_game_id=f"G{i:03}",
            game_date=date(2025, 11, i),
            opponent_name="Historical Opponent",
            venue_type=VenueType.HOME,
            opponent_score=0,
            workbook_team_score=2,
            workbook_result="WIN",
        )
        for i in range(1, 16)
    ]

    player_stats = [
        HistoricalPlayerStatsRow(
            workbook_game_id=f"G{i:03}",
            player_name="Kevin Lans",
            three_point_attempts=0,
            three_point_makes=0,
            two_point_attempts=1,
            two_point_makes=1,
            free_throw_attempts=0,
            free_throw_makes=0,
            turnovers=0,
            assists=0,
            offensive_rebounds=0,
            defensive_rebounds=0,
            steals=0,
            deflections=0,
            personal_fouls=0,
        )
        for i in range(1, 16)
    ]

    data = HistoricalWorkbookData(
        roster=[
            HistoricalRosterRow(
                full_name="Kevin Lans",
                jersey_number=2,
                position="Guard",
                grade_level="11",
            )
        ],
        games=games,
        player_stats=player_stats,
    )

    import_historical_data(
        db_session,
        data,
    )

    with pytest.raises(
        HistoricalImportAlreadyExistsError,
        match="already exists",
    ):
        import_historical_data(
            db_session,
            data,
        )

    seasons = list(
        db_session.scalars(
            select(Season)
        ).all()
    )

    games_in_database = list(
        db_session.scalars(
            select(Game)
        ).all()
    )

    stats_in_database = list(
        db_session.scalars(
            select(PlayerGameStats)
        ).all()
    )

    assert len(seasons) == 1
    assert len(games_in_database) == 15
    assert len(stats_in_database) == 15